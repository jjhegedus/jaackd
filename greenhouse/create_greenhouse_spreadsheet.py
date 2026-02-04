import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Create a new workbook and add sheets
wb = Workbook()

# Rename default sheet to "Inputs"
ws_inputs = wb.active
ws_inputs.title = "Inputs"

# Create "Calculations" sheet
ws_calc = wb.create_sheet(title="Calculations")

# -----------------------------
# Sheet 1: Inputs
# -----------------------------
inputs_headers = ["Parameter", "Value", "Unit", "Description"]
inputs_data = [
    ["Greenhouse Length", 40, "ft", "East–west dimension"],
    ["Greenhouse Width", 24, "ft", "North–south dimension"],
    ["Greenhouse Height", 19.33, "ft", "Height up to just below the pole building’s eve (19′4″ ≈ 19.33 ft)"],
    ["U-value (Polycarbonate)", 2.5, "W/m²·K", "Thermal transmittance of 8 mm twin wall polycarbonate"],
    ["Transmittance (τ)", 0.75, "", "Solar transmittance for the glazing"],
    ["Solar Irradiance (I)", 800, "W/m²", "Example value for full sun (adjust as needed)"],
    ["Solar Angle (θ)", 30, "degrees", "Angle of incidence for the south wall"],
    ["North Wall Temperature (T_north)", 65, "°F", "Fixed temperature provided by the pole building (daytime)"],
    ["Ambient Temperature (T_ambient)", 50, "°F", "Outside air temperature (example value)"],
    ["Initial Internal Temperature (T_inside)", 60, "°F", "Starting estimate for internal temperature"],
    ["External Convection Coefficient (h_ext)", 10, "W/m²·K", "Estimate for wind effects in Indiana"],
    ["Thermal Capacitance (C)", 100000, "J/K", "Effective thermal mass of the greenhouse (estimate)"]
]

# Write headers
for col, header in enumerate(inputs_headers, start=1):
    ws_inputs.cell(row=1, column=col, value=header)

# Write data rows
for row_index, row_data in enumerate(inputs_data, start=2):
    for col_index, value in enumerate(row_data, start=1):
        ws_inputs.cell(row=row_index, column=col_index, value=value)

# Adjust column widths for Inputs sheet
for col in range(1, len(inputs_headers) + 1):
    ws_inputs.column_dimensions[get_column_letter(col)].width = 25

# -----------------------------
# Sheet 2: Calculations
# -----------------------------
calc_headers = ["Calculation", "Formula/Value", "Unit", "Notes"]
calc_data = [
    # Convert dimensions (ft² to m² conversion factor)
    ["Conversion Factor", "0.092903", "", "1 ft² = 0.092903 m²"],
    
    # South Wall Area
    ["South Wall Area (ft²)", "=Inputs!B2*Inputs!B3", "ft²", "Length * Height (assuming south wall uses full height)"],
    ["South Wall Area (m²)", "=B2*B1", "m²", "Convert ft² to m² (use conversion factor)"],
    
    # Solar Gain Calculations
    ["Cosine Correction", "=COS(RADIANS(Inputs!B7))", "", "Solar angle in radians"],
    ["Effective Solar Irradiance (I_eff)", "=Inputs!B6*B4", "W/m²", "I * cosine correction"],
    ["Solar Gain (Q_solar)", "=Inputs!B5*B5*B2", "W", "τ * I_eff * South Wall Area (m²)"],
    
    # Conduction/Convection (example for south wall)
    ["Conduction Loss (South Wall)", "=Inputs!B4*B2*(Inputs!B10 - Inputs!B9)", "W", "U-value * Area (m²) * (T_inside - T_ambient)"],
    
    # North Wall Exchange (attached to pole building)
    ["North Wall Heat Exchange", "=Inputs!B4*B2*(Inputs!B8 - Inputs!B10)", "W", "U-value * North Wall Area (m²) * (T_north - T_inside)"]
]

# Write headers for Calculations sheet
for col, header in enumerate(calc_headers, start=1):
    ws_calc.cell(row=1, column=col, value=header)

# Write data rows for Calculations sheet
for row_index, row_data in enumerate(calc_data, start=2):
    for col_index, value in enumerate(row_data, start=1):
        ws_calc.cell(row=row_index, column=col_index, value=value)

# Adjust column widths for Calculations sheet
for col in range(1, len(calc_headers) + 1):
    ws_calc.column_dimensions[get_column_letter(col)].width = 40

# Save the workbook to a file
output_filename = "GreenhouseThermalModel.xlsx"
wb.save(output_filename)
print(f"Spreadsheet '{output_filename}' has been created.")
